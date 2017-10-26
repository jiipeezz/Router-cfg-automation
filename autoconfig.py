import paramiko
import sys
import socket
import os
import time
import openpyxl

def get_serial():
	cmd = "status -v sys |grep \"Serial Number\" |awk '{print $4}'"
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	outp = ssh_stdout.readlines()
	serial = outp[0].strip()
	return serial

def get_mac():
	cmd = "ifconfig eth0 |grep \"HWaddr\" |awk '{print $5}'"
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	outp = ssh_stdout.readlines()
	mac = outp[0].strip()
	return mac

def restore_cfg(restore_file):
	orig = restore_file
	dest = "/root/" + restore_file
	cmd = "restore " + dest
	sftp = ssh.open_sftp()
	sftp.put(orig, dest)
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	success = "Configuration successfully updated."
	outp = ssh_stdout.readlines()
	status = outp[0].strip()
	if status == success:
		status_msg = "OK"
	else:
		status_msg = "FAILED"
	return status_msg

def change_snmp(serial):
	cmd = "sed -i 's/SNMP_NAME=.*/SNMP_NAME=" + str(serial) + "/' /etc/settings.snmp"
	check = "sed -n 's/SNMP_NAME=//p' /etc/settings.snmp"
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	time.sleep(2)
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(check)
	outp = ssh_stdout.readlines()
	serialcheck = outp[0].strip()
	if str(serialcheck) == str(serial):
		state = "OK"
	else:
		state = "FAILED"
	return state

def add_um(user_m, m_name):
	orig = user_m
	dest = "/opt/" + user_m
	cmd = "tar -xzf " + dest + " -C /opt/"
	check = "if [ -d \"/opt/" + m_name + "\" ];then echo OK;else echo NOT;fi"
	sftp = ssh.open_sftp()
	sftp.put(orig, dest)
	time.sleep(2)
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	time.sleep(1)
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(check)
	outp = ssh_stdout.readlines()
	status = outp[0].strip()
	if status == "OK":
		status_msg = "OK"
	else:
		status_msg = "FAILED"
	return status_msg

def change_pw(passwd):
	cmd = "echo 'root:"+ str(passwd) + "' |chpasswd"
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	outp = ssh_stdout.readlines()
	status = outp[0].strip()
	success = "Password for 'root' changed"
	if status == success:
		status_msg = "OK"
	else:
		status_msg = "FAILED"
	return status_msg

def get_backup(filename):
	bu_file = "bckup" + filename
	cmd = "backup > " + bu_file
	orig = "/root/" + bu_file
	ssh_stdin, ssh_stdout, ssh_stderr = ssh.exec_command(cmd)
	time.sleep(2)
	sftp = ssh.open_sftp()
	sftp.get(orig, bu_file)

def update_excel(ser, maci):
	#opening excel workbook
	wb = openpyxl.load_workbook(filename = excelfile)

	#opening the first sheet
	sheets = wb.sheetnames
	ws = wb[sheets[0]]

	count = 1
	i = 0

	letters = ["A", "B", "C", "D", "E", "F", "J"]
	while 1:
		total = letters[i] + str(count)
		valuecheck = ws[total].value
		if not valuecheck:
			if letters[i] == "A":
				ws[total] = fullip
				print("VPN IP updated to excel")
			elif letters[i] == "B":
				ws[total] = mask
				print("Network mask updated to excel")
			elif letters[i] == "C":
				ws[total] = ser
				print("Serial number updated to excel")
			elif letters[i] == "D":
				ws[total] = maci
				print("MAC address updated to excel")
			elif letters[i] == "E":
				ws[total] = model
				print("Model updated to excel")
			elif letters[i] == "F":
				ws[total] = date
				print("Date updated to excel")
			elif letters[i] == "J":
				ws[total] = reference
				print("Reference updated to excel")
				break
			count = 1
			i += 1
		count += 1

	wb.save(filename = excelfile)

try:
	model = sys.argv[1]
	reference = sys.argv[2]
	vpnip = sys.argv[3]
except IndexError:
	print("Usage: python autoconfig.py <model> <reference> <first 3 of VPNIP/xxx.xxx.xxx>")
	sys.exit()

router_dflt_ip = "192.168.9.1"
uname = "root"
passwd = "<password>"
new_passwd = "<password>"
user_m1 = "pinger.v3.tgz"
user_m1_name = "pinger"
user_m2 = "hmpclient.v2.tgz"
user_m2_name = "hmpclient"
date = time.strftime("%d/%m/%Y")
excelfile = "customer_router_information.xlsx"
mask = "255.255.255.0"
restore_file = "testcfg_" + vpnip + ".cfg"
fullip = vpnip + ".1"

#checking if necessary files exist in current working directory
if not os.path.exists(user_m1):
	print("Unable to find user module " + user_m1 + " in current working directory.")
	sys.exit()
	
if not os.path.exists(user_m2):
	print("Unable to find user module " + user_m2 + " in current working directory.")
	sys.exit()
	
if not os.path.exists(restore_file):
	print("Unable to find cfg file " + restore_file + " in current working directory.")
	sys.exit()

if not os.path.exists(excelfile):
	print("Unable to find excel file " + excelfile + " in current working directory.")
	sys.exit()

try:
	ssh = paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	ssh.connect(router_dflt_ip, username=uname, password=passwd, timeout=15)
#catching misconfigured network settings
except socket.timeout:
	print("Connection timed out. Check your network settings")
	sys.exit()
#catching incorrect credentials
except paramiko.ssh_exception.AuthenticationException:
	print("Authentication error. Check your credentials")
	sys.exit()

serial = get_serial()
mac = get_mac()

print("Serial number: " + str(serial))
print("MAC address: " + mac + "\n")
print("Starting configuration daemon...")

print("Sending configuration file...")
while True:
	restore_status = restore_cfg(restore_file)
	if restore_status == "OK":
		break

print("Changing SNMP name...")
while True:
	snmp_status = change_snmp(serial)
	if snmp_status == "OK":
		break

print("Adding user module " + user_m1 + "...")
while True:
	userm_status1 = add_um(user_m1, user_m1_name)
	if userm_status1 == "OK":
		break

print("Adding user module " + user_m2 + "...")
while True:
	userm_status2 = add_um(user_m2, user_m2_name)
	if userm_status2 == "OK":
		break

print("Changing root's password...\n")
while True:
	pw_status = change_pw(new_passwd)
	if pw_status == "OK":
		break

print("Verification summary")
print("----------------------------------")
print("|Configuration file restored: " + restore_status + " |")
print("|SNMP Changed: " + snmp_status + "                |")
print("|User module " + user_m1 + ": " + userm_status1 + "   |")
print("|User module " + user_m2 + ": " + userm_status2 + "|")
print("|Password changed: " + pw_status + "            |")
print("----------------------------------\n")
print("Saving backup configuration file as bckup" + restore_file + "...\n")
get_backup(restore_file)
update_excel(serial, mac)
print("\nDone. The router can be unplugged now!")

ssh.close()
